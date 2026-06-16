import streamlit as st
import google.generativeai as genai
import pandas as pd
import json
import time
import os
import tempfile
import io
import datetime
import calendar
import re
import openpyxl
from openpyxl import load_workbook
from copy import copy
 
# ==========================================
# 1. API 설정 및 기본 함수
# ==========================================
API_KEY = st.secrets["GEMINI_API_KEY"]
genai.configure(api_key=API_KEY)
 
def get_best_model():
    try:
        valid_models = []
        for m in genai.list_models():
            if 'generateContent' in m.supported_generation_methods:
                valid_models.append(m.name)
        for m in valid_models:
            if 'flash' in m.lower():
                return m
        return valid_models[0] if valid_models else "models/gemini-pro"
    except Exception:
        return "models/gemini-pro"
 
def format_company_name(name):
    if not name:
        return ""
    name = re.sub(r'주식회사\s*', '㈜', name)
    name = re.sub(r'\s*주식회사', '㈜', name)
    name = re.sub(r'\(\s*주\s*\)\s*', '㈜', name)
    name = re.sub(r'\s*\(\s*주\s*\)', '㈜', name)
    name = name.replace('㈜㈜', '㈜')
    return name.strip()
 
def format_region(region_str):
    if not region_str:
        return "", ""
    mapping_rules = [
        (["서울"], "02 서울"),
        (["부산"], "051 부산"),
        (["대구"], "053 대구"),
        (["인천"], "032 인천"),
        (["광주"], "062 광주"),
        (["대전"], "042 대전"),
        (["울산"], "052 울산"),
        (["세종"], "044 세종"),
        (["경기"], "031 경기"),
        (["강원"], "033 강원"),
        (["충북", "충청북도"], "043 충북"),
        (["충남", "충청남도"], "041 충남"),
        (["전북", "전라북도", "전북특별자치도"], "063 전북"),
        (["전남", "전라남도"], "061 전남"),
        (["경북", "경상북도"], "054 경북"),
        (["경남", "경상남도"], "055 경남"),
        (["제주"], "064 제주"),
    ]
    for keys, value in mapping_rules:
        for k in keys:
            if k in region_str:
                return "국내", value
    return "", region_str
 
def format_currency(value):
    if not value:
        return ""
    try:
        clean_num = re.sub(r'[^\d]', '', str(value))
        if clean_num:
            return f"{int(clean_num):,}"
    except Exception:
        pass
    return str(value)
 
def add_months(sourcedate, months):
    month = sourcedate.month - 1 + months
    year = sourcedate.year + month // 12
    month = month % 12 + 1
    day = min(sourcedate.day, calendar.monthrange(year, month)[1])
    return datetime.date(year, month, day)
 
def calculate_exact_period(start_date_str, period_str):
    try:
        start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
        year_match = re.search(r"(\d+)\s*년", period_str)
        month_match = re.search(r"(\d+)\s*(?:개월|월)", period_str)
        if year_match:
            end_date = add_months(start_date, int(year_match.group(1)) * 12) - datetime.timedelta(days=1)
            return f"{start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}"
        elif month_match:
            end_date = add_months(start_date, int(month_match.group(1))) - datetime.timedelta(days=1)
            return f"{start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}"
    except Exception:
        pass
    return period_str
 
# ==========================================
# 2. Gemini 추출 함수 (계약 정보)
# ==========================================
def extract_with_gemini(contract_path, biz_reg_path, info_path, model_name):
    model = genai.GenerativeModel(model_name)
    uploaded_files = []
    try:
        c_file = genai.upload_file(path=contract_path)
        uploaded_files.append(c_file)
        docs_to_analyze = [c_file]
        time.sleep(2)
        if biz_reg_path:
            b_file = genai.upload_file(path=biz_reg_path)
            uploaded_files.append(b_file)
            docs_to_analyze.append(b_file)
            time.sleep(2)
        if info_path:
            i_file = genai.upload_file(path=info_path)
            uploaded_files.append(i_file)
            docs_to_analyze.append(i_file)
            time.sleep(2)
 
        prompt = """
        첨부된 문서들을 종합 분석하여 아래 항목의 정보를 추출해 줘.
        반드시 마크다운 기호 없이 순수 JSON 형식으로만 답변해야 해. 정보가 없으면 빈 문자열("") 입력.
        {
            "1. 기술이전계약일": "YYYY-MM-DD",
            "2. 회사명": "계약 상대방 업체명",
            "3. 회사 주소": "괄호 안의 건물명 제외 지번/도로명까지",
            "4. 회사 대표명": "",
            "5. 사업자등록번호": "000-00-00000",
            "6. 지역구분": "회사 주소의 시/도 단위",
            "7. 회사 업무담당자 성명": "",
            "8. 회사 업무 담당자 이메일": "",
            "9. 회사 업무 담당자 번호": "010-0000-0000",
            "10. 기술이전계약명": "계약명 또는 발명의 명칭",
            "11. 기술이전책임자명": "주발명자명 기재",
            "12. 학과": "소속(단과대학) 또는 소속 학과",
            "13. 기술유형": "'특허', '노하우', '자문', '저작권' 중 하나",
            "14. 거래유형": "'독점 통상실시권', '비독점 통상실시권', '전용실시권', '특허양도' 중 하나",
            "15. 계약기간": "계약서에 명시된 계약기간. 양도계약이면 빈 문자열",
            "16. 기술료 유형": "'정액기술료', '경상기술료', '혼합형(정액+경상)' 중 하나",
            "17. 총 정액기술료(단위: 원)": "아라비아 숫자로 변환 (콤마 제외)",
            "18. 정액기술료 납부방법": "일시불 또는 분할 스케줄 상세 기재. 없으면 해당없음",
            "19. 경상기술료(Running Royalty) 조건": "조건 상세 기재. 없으면 해당없음",
            "20. 학교 업무담당자 성명": "산학협력단 측 담당자 성명",
            "21-1. 특허출원번호": "",
            "21-2. 특허등록번호": "",
            "22. 주발명자 전화번호": "",
            "23. 주발명자 이메일": "",
            "24. 연구과제명": "",
            "25. 대사업명": "",
            "26. 중사업명": "",
            "27. 지원기관과제번호": "",
            "28. 연구협약일": "",
            "29. 정부출연금": "",
            "30. 총연구비": "",
            "31. 총연구기간": "",
            "32. 연구책임자": "",
            "33. 수납상황": "",
            "34. 업체 비용담당자 성명": "",
            "35. 업체 비용담당자 부서": "",
            "36. 업체 비용담당자 직급": "",
            "37. 업체 비용담당자 전화번호": "",
            "38. 업체 비용담당자 이메일": "",
            "39. 기술분야(6T)": "기술이전계약명(기술명)을 보고 아래 목록 중 가장 적합한 1개만 선택해서 기재. 선택지: IT, BT, NT, CT, ET, ST, 기타. (IT=정보통신기술, BT=생명공학기술, NT=나노기술, CT=문화기술, ET=환경기술, ST=우주항공기술)",
            "40. 기술분류": "기술이전계약명(기술명)을 보고 아래 목록 중 가장 적합한 1개만 선택해서 기재. 선택지: 수학, 물리학, 화학, 지구과학(지구·대기·해양·천문), 생명과학, 농림수산식품, 보건의료, 기계, 재료, 화공, 전기/전자, 정보통신, 에너지/자원, 원자력, 환경, 건설/교통, 역사/고고학, 철학/종교, 언어, 문학, 문화/예술/체육, 법, 정치/행정, 경제/경영, 사회/인류/복지/여성, 생활, 지리/지역/관광, 심리, 교육, 미디어/커뮤니케이션/문헌정보, 뇌과학, 인지/감성과학, 과학기술과 인문사회, 인력 및 인프라, 기타",
            "41. 기관유형": "사업자등록증의 회사 규모/유형을 보고 아래 목록 중 가장 적합한 1개만 선택. 선택지: 대기업, 중견기업, 중소기업(일반), 중소기업(벤처), 개인, 국공립대학, 사립대학, 국공립시험연구기관, 정부출연연구기관, 특정연구기관, 전문생산기술연구소, 기술거래기관, 기타 비영리 법인 및 단체, 해외, 기타 정부산하기관, 공공기관, 공기업, 기타. 사업자등록증에 중소기업 확인서나 벤처기업 확인서 내용이 있으면 참고. 정보가 부족하면 빈 문자열",
            "42. 업종유형": "사업자등록증의 업태와 종목을 바탕으로 한국표준산업분류(KSIC) 코드와 업종명을 추출. 형식 예시: (C26422)이동전화기 제조업 / (J58222)응용 소프트웨어 개발 및 공급업. 코드를 알 수 없으면 업태와 종목만 기재. 정보가 없으면 빈 문자열"
        }
        """
        docs_to_analyze.append(prompt)
 
        max_retries = 2
        for attempt in range(max_retries):
            try:
                response = model.generate_content(docs_to_analyze, request_options={"timeout": 600})
                break
            except Exception as e:
                if '429' in str(e) or 'exceeded' in str(e):
                    if attempt < max_retries - 1:
                        time.sleep(45)
                        continue
                raise e
 
        result_text = response.text.strip()
        md_json = chr(96)*3 + "json"
        md_end = chr(96)*3
        if result_text.startswith(md_json):
            result_text = result_text[7:]
        elif result_text.startswith(md_end):
            result_text = result_text[3:]
        if result_text.endswith(md_end):
            result_text = result_text[:-3]
 
        extracted_data = json.loads(result_text.strip())
 
        if "2. 회사명" in extracted_data:
            extracted_data["2. 회사명"] = format_company_name(extracted_data["2. 회사명"])
 
        deal_type = extracted_data.get("14. 거래유형", "")
        start_date = extracted_data.get("1. 기술이전계약일", "")
        raw_period = extracted_data.get("15. 계약기간", "")
 
        if "양도" in deal_type:
            extracted_data["15. 계약기간"] = "특허존속기간 만료일까지"
        elif start_date and raw_period:
            if "~" not in raw_period and ("년" in raw_period or "개월" in raw_period or "월" in raw_period):
                extracted_data["15. 계약기간"] = calculate_exact_period(start_date, raw_period)
 
        return extracted_data
 
    except Exception as e:
        return {"10. 기술이전계약명": f"오류 발생: {e}"}
    finally:
        for f in uploaded_files:
            try:
                genai.delete_file(f.name)
            except:
                pass
 
# ==========================================
# 3. Gemini 추출 함수 (기술료 분배)
# ==========================================
def extract_distribution_with_gemini(dist_path, model_name):
    model = genai.GenerativeModel(model_name)
    uploaded_files = []
    try:
        d_file = genai.upload_file(path=dist_path)
        uploaded_files.append(d_file)
        time.sleep(2)
 
        prompt = """
        첨부된 기술료 분배 문서를 분석하여 아래 항목을 추출해 줘.
        반드시 마크다운 기호 없이 순수 JSON 형식으로만 답변해야 해. 정보가 없으면 빈 문자열("") 입력.
        연번은 문서 제목에 포함된 연번(예: 2026-009, 2025-086 등)을 추출해.
        입금액이 여러 번인 경우 모든 입금일과 입금액을 합산하여 기재해.
 
        {
            "연번": "문서 제목의 연번 (예: 2026-009)",
            "입금일": "가장 마지막 입금일 YYYY-MM-DD 형식. 여러 건이면 쉼표 구분",
            "입금액합계": "총 입금액 숫자만 (콤마 제외, 부가세 포함 금액)",
            "분배기준액": "C=A-B 분배기준액 숫자만 (콤마 제외)",
            "발명자보상금": "발명자(저작자) 보상금 숫자만 (콤마 제외)",
            "지식재산권비용": "기술이전사업화경비 지식재산권 출원등록유지 금액 숫자만 (없으면 0)",
            "성과활용기여자보상금": "성과활용 기여자 보상금 숫자만 (없으면 0)",
            "연구개발재투자": "연구개발 재투자/기관운영경비 금액 숫자만 (없으면 0)",
            "중개수수료": "중개수수료 금액 숫자만 (없으면 0)",
            "특허비용공제": "특허비용공제 금액 숫자만 (없으면 0)",
            "산학협력단분배액": "산학협력단 직접 분배액 숫자만 (자문/저작권 건에서 존재. 없으면 0)",
            "지정기관분배액": "지정기관(교양교육원 등) 분배액 숫자만 (없으면 0)",
            "분배일": "문서 작성일 또는 결재일 YYYY-MM-DD 형식"
        }
        """
 
        max_retries = 2
        for attempt in range(max_retries):
            try:
                response = model.generate_content([d_file, prompt], request_options={"timeout": 300})
                break
            except Exception as e:
                if '429' in str(e) or 'exceeded' in str(e):
                    if attempt < max_retries - 1:
                        time.sleep(45)
                        continue
                raise e
 
        result_text = response.text.strip()
        md_json = chr(96)*3 + "json"
        md_end = chr(96)*3
        if result_text.startswith(md_json):
            result_text = result_text[7:]
        elif result_text.startswith(md_end):
            result_text = result_text[3:]
        if result_text.endswith(md_end):
            result_text = result_text[:-3]
 
        return json.loads(result_text.strip())
 
    except Exception as e:
        return {"연번": "", "오류": str(e)}
    finally:
        for f in uploaded_files:
            try:
                genai.delete_file(f.name)
            except:
                pass
 
# ==========================================
# 4. 총정리파일 행 추가 함수
# ==========================================
def append_row_to_master(master_path, extracted_data, target_year):
    wb = load_workbook(master_path)
    ws = wb['내역']
 
    # 마지막 실제 데이터 행 찾기
    # A열이 YYYY-NNN 연번 패턴이거나 B열이 datetime인 행 중 마지막
    last_row = 1
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        a_val = str(row[0]) if row[0] is not None else ""
        b_is_date = isinstance(row[1], datetime.datetime)
        a_is_serial = bool(re.match(r'^\d{4}-\d{3}', a_val))
        if a_is_serial or b_is_date:
            last_row = i + 2
 
    new_row = last_row + 1
 
    # 연번 직접 계산 (openpyxl은 수식 실행 불가 → Python에서 직접 계산 후 문자열로 저장)
    last_serial = ws.cell(row=last_row, column=1).value
    serial_match = re.search(r'(\d{4})-(\d+)', str(last_serial)) if last_serial else None
    if serial_match:
        new_serial = f'{serial_match.group(1)}-{int(serial_match.group(2))+1:03d}'
    else:
        new_serial = f'{target_year}-001'
    ws.cell(row=new_row, column=1).value = new_serial
 
    # 컬럼 헤더 → 열 번호 매핑
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column
 
    def set_col(header_keyword, value):
        for h, col in headers.items():
            if header_keyword in h:
                ws.cell(row=new_row, column=col).value = value
                return
 
    d = extracted_data
 
    # 계약 기본 정보
    contract_date = d.get("1. 기술이전계약일", "")
    if contract_date:
        try:
            ws.cell(row=new_row, column=2).value = datetime.datetime.strptime(contract_date, "%Y-%m-%d").date()
        except:
            ws.cell(row=new_row, column=2).value = contract_date
 
    ws.cell(row=new_row, column=3).value = "부산대학교 산학협력단"
    ws.cell(row=new_row, column=4).value = d.get("2. 회사명", "")
    ws.cell(row=new_row, column=5).value = d.get("41. 기관유형", "")  # 5.기관유형
    ws.cell(row=new_row, column=6).value = d.get("42. 업종유형", "")  # 6.업종유형
 
    raw_region = d.get("6. 지역구분", "")
    dom_ovs, formatted_region = format_region(raw_region)
    set_col("7.국내", dom_ovs)
    set_col("9.국내지역", formatted_region)
    set_col("10. 사업자", d.get("5. 사업자등록번호", ""))
    set_col("11. 대표주소", d.get("3. 회사 주소", ""))
    set_col("13. 대표자성명", d.get("4. 회사 대표명", ""))
    set_col("15. 기술이전담당", d.get("7. 회사 업무담당자 성명", ""))
    set_col("19.이메일", d.get("8. 회사 업무 담당자 이메일", ""))
    set_col("18.핸드폰", d.get("9. 회사 업무 담당자 번호", ""))
    set_col("27.기술명", d.get("10. 기술이전계약명", ""))
    set_col("28.주발명자", d.get("11. 기술이전책임자명", ""))
    set_col("29.소속", d.get("12. 학과", ""))
    set_col("34.기술유형", d.get("13. 기술유형", ""))
    set_col("39.기술분야", d.get("39. 기술분야(6T)", ""))
    set_col("40.기술분류", d.get("40. 기술분류", ""))
    set_col("41.거래유형", d.get("14. 거래유형", ""))
    set_col("42.계약기간", d.get("15. 계약기간", ""))
    set_col("46.기술료 수취유형", d.get("16. 기술료 유형", ""))
    set_col("50.총 기술료", format_currency(d.get("17. 총 정액기술료(단위: 원)", "")))
    set_col("납부기한", d.get("18. 정액기술료 납부방법", ""))
    set_col("48.경상기술료", d.get("19. 경상기술료(Running Royalty) 조건", ""))
 
    # 비용담당자
    set_col("20.담당자명", d.get("34. 업체 비용담당자 성명", ""))
    set_col("21.담당부서", d.get("35. 업체 비용담당자 부서", ""))
    set_col("22.직급", d.get("36. 업체 비용담당자 직급", ""))
    set_col("23.핸드폰/", d.get("37. 업체 비용담당자 전화번호", ""))
    set_col("24.이메일", d.get("38. 업체 비용담당자 이메일", ""))
 
    # 지식재산권 번호/상태
    reg_no = d.get("21-2. 특허등록번호", "").strip()
    app_no = d.get("21-1. 특허출원번호", "").strip()
    if reg_no:
        set_col("35.지식재산권 번호", reg_no)
        set_col("36.상태", "등록")
    elif app_no:
        set_col("35.지식재산권 번호", app_no)
        set_col("36.상태", "출원")
 
    # 연구과제 정보
    set_col("53.연구과제명", d.get("24. 연구과제명", ""))
    set_col("60.대사업명", d.get("25. 대사업명", ""))
    set_col("61.중사업명", d.get("26. 중사업명", ""))
    set_col("62.지원기관과제번호", d.get("27. 지원기관과제번호", ""))
    set_col("59. 협약일", d.get("28. 연구협약일", ""))
    set_col("67.정부출연금", format_currency(d.get("29. 정부출연금", "")))
    set_col("57.총연구비", format_currency(d.get("30. 총연구비", "")))
    set_col("58.총연구기간", d.get("31. 총연구기간", ""))
    set_col("64.연구책임자", d.get("32. 연구책임자", ""))
    set_col("91.수납상황", d.get("33. 수납상황", ""))
 
    # 담당자
    set_col("담당자", d.get("20. 학교 업무담당자 성명", ""))
 
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
 
# ==========================================
# 5. 총정리파일 분배 업데이트 함수
# ==========================================
def update_distribution_in_master(master_path, dist_data_list):
    wb = load_workbook(master_path)
    ws = wb['내역']
 
    # 헤더 → 열 번호 매핑
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column
 
    def get_col(keyword):
        for h, col in headers.items():
            if keyword in h:
                return col
        return None
 
    # 연번 → 행 번호 목록 매핑 (A열 기준, 동일 연번 여러 행 수집)
    serial_to_rows = {}
    for row in ws.iter_rows(min_row=2):
        val = row[0].value
        if val and isinstance(val, str) and re.match(r'\d{4}-\d{3}', val):
            serial = val.strip()
            if serial not in serial_to_rows:
                serial_to_rows[serial] = []
            serial_to_rows[serial].append(row[0].row)
 
    results = []
    for dist_data in dist_data_list:
        serial = dist_data.get("연번", "").strip()
        if not serial:
            results.append({"연번": "연번 추출 실패", "상태": "❌ 실패"})
            continue
 
        if serial not in serial_to_rows:
            results.append({"연번": serial, "상태": f"❌ 총정리파일에서 '{serial}' 행을 찾지 못함"})
            continue
 
        all_rows = serial_to_rows[serial]
        # 동일 연번 중 마지막 행에만 분배 데이터 입력
        target_row = all_rows[-1]
 
        def safe_int(val):
            try:
                return int(re.sub(r'[^\d]', '', str(val))) if val else 0
            except:
                return 0
 
        # 70.입금일 / 73.현금입금액 / 82.분배일 는 입력 제외
        col = get_col("83.제반비용\n(특허비용)")
        if col:
            ws.cell(row=target_row, column=col).value = safe_int(dist_data.get("특허비용공제"))
 
        col = get_col("84.제반비용\n(중개수수료")
        if col:
            ws.cell(row=target_row, column=col).value = safe_int(dist_data.get("중개수수료"))
 
        col = get_col("86.발명자")
        if col:
            ws.cell(row=target_row, column=col).value = safe_int(dist_data.get("발명자보상금"))
 
        col = get_col("88.산학협력단")
        if col:
            ws.cell(row=target_row, column=col).value = safe_int(dist_data.get("산학협력단분배액"))
 
        col = get_col("88-1.")
        if col:
            ws.cell(row=target_row, column=col).value = safe_int(dist_data.get("지식재산권비용"))
 
        col = get_col("88-2.")
        if col:
            ws.cell(row=target_row, column=col).value = safe_int(dist_data.get("성과활용기여자보상금"))
 
        col = get_col("88-3.")
        if col:
            ws.cell(row=target_row, column=col).value = safe_int(dist_data.get("연구개발재투자"))
 
        row_info = f"총 {len(all_rows)}행 중 마지막 행({target_row})에 입력"
        results.append({"연번": serial, "상태": f"✅ {row_info}"})
 
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue(), results
 
# ==========================================
# 6. Streamlit UI
# ==========================================
st.set_page_config(page_title="기술이전 통합 관리 시스템", page_icon="📑", layout="wide")
st.title("📑 기술이전 통합 자동화 시스템")
 
tab1, tab2 = st.tabs(["📄 1단계: 계약 추출 → 총정리파일 추가", "💰 2단계: 기술료 분배 업데이트"])
 
# ==========================================
# TAB 1: 계약 추출
# ==========================================
with tab1:
    st.markdown("""
    계약서·사업자등록증·기술이전정보 PDF와 **기존 기술이전총정리파일**을 함께 업로드하면,
    AI가 내용을 추출하여 총정리파일 마지막 행에 자동으로 추가합니다.
    """)
    st.warning("""
    ⚠️ **여러 건 업로드 시 반드시 아래 순서를 지켜주세요!**
    
    계약서 / 사업자등록증 / 기술이전정보는 **같은 순서**로 업로드해야 올바르게 매칭됩니다.
    
    예) A건 계약서 → B건 계약서 순으로 올렸다면, 사업자등록증도 A건 → B건, 기술이전정보도 A건 → B건 순으로 올려주세요.
    """)
 
    col1, col2 = st.columns(2)
    with col1:
        master_file_tab1 = st.file_uploader(
            "📊 기술이전 총정리파일 (.xlsx)", type=['xlsx'], key="master1"
        )
        target_year = st.number_input(
            "📅 연번 연도", min_value=2000, value=datetime.date.today().year, step=1
        )
    with col2:
        st.info("💡 연번은 총정리파일의 마지막 연번에서 자동으로 +1됩니다.")
 
    st.markdown("---")
    col_a, col_b, col_c = st.columns(3)
    with col_a:
        contract_files = st.file_uploader(
            "1. 기술이전계약서 (PDF) 📄\n여러 건은 순서대로 업로드", type=['pdf'], accept_multiple_files=True, key="contract"
        )
    with col_b:
        biz_files = st.file_uploader(
            "2. 사업자등록증 (PDF) 🏢\n계약서와 동일한 순서로 업로드", type=['pdf'], accept_multiple_files=True, key="biz"
        )
    with col_c:
        info_files = st.file_uploader(
            "3. 기술이전 정보 (PDF) 💡\n계약서와 동일한 순서로 업로드", type=['pdf'], accept_multiple_files=True, key="info"
        )
 
    if st.button("🚀 추출 시작 → 총정리파일에 추가", use_container_width=True, key="btn1"):
        if not master_file_tab1:
            st.error("⚠️ 기술이전 총정리파일을 업로드해 주세요!")
        elif not contract_files:
            st.error("⚠️ 최소 1개 이상의 계약서 PDF를 업로드해 주세요!")
        else:
            model_name = get_best_model()
 
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_master:
                tmp_master.write(master_file_tab1.read())
                master_path = tmp_master.name
 
            all_extracted = []
            progress_bar = st.progress(0)
            status_text = st.empty()
 
            for i, c_file in enumerate(contract_files):
                status_text.info(f"⏳ [{i+1}/{len(contract_files)}] '{c_file.name}' 분석 중...")
 
                with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_c:
                    tmp_c.write(c_file.read())
                    c_path = tmp_c.name
 
                b_path = ""
                if biz_files and len(biz_files) > i:
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_b:
                        tmp_b.write(biz_files[i].read())
                        b_path = tmp_b.name
 
                i_path = ""
                if info_files and len(info_files) > i:
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_i:
                        tmp_i.write(info_files[i].read())
                        i_path = tmp_i.name
 
                data = extract_with_gemini(c_path, b_path, i_path, model_name)
                all_extracted.append(data)
 
                os.remove(c_path)
                if b_path: os.remove(b_path)
                if i_path: os.remove(i_path)
 
                progress_bar.progress((i + 1) / len(contract_files))
 
            status_text.info("📝 총정리파일에 데이터 추가 중...")
 
            current_master_path = master_path
            for data in all_extracted:
                updated_bytes = append_row_to_master(current_master_path, data, target_year)
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_out:
                    tmp_out.write(updated_bytes)
                    current_master_path = tmp_out.name
 
            with open(current_master_path, 'rb') as f:
                final_bytes = f.read()
 
            os.remove(master_path)
 
            status_text.success(f"🎉 {len(all_extracted)}건 추출 완료! 총정리파일에 추가되었습니다.")
 
            # 미리보기
            preview_data = []
            for d in all_extracted:
                preview_data.append({
                    "기술이전계약명": d.get("10. 기술이전계약명", ""),
                    "회사명": d.get("2. 회사명", ""),
                    "계약일": d.get("1. 기술이전계약일", ""),
                    "거래유형": d.get("14. 거래유형", ""),
                    "기술유형": d.get("13. 기술유형", ""),
                })
            st.dataframe(pd.DataFrame(preview_data), use_container_width=True)
 
            today = datetime.date.today().strftime("%Y%m%d")
            st.download_button(
                label="📥 업데이트된 기술이전총정리파일 다운로드",
                data=final_bytes,
                file_name=f"기술이전총정리_{today}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
 
# ==========================================
# TAB 2: 기술료 분배 업데이트
# ==========================================
with tab2:
    st.markdown("""
    기술료 입금 완료 후, **분배 PDF**와 **기술이전총정리파일**을 업로드하면
    연번을 기준으로 해당 행의 분배 관련 열을 자동으로 업데이트합니다.
    
    - 분배 PDF는 여러 건을 한꺼번에 업로드 가능합니다.
    - 업데이트 대상 열: `70.입금일`, `73.현금입금액`, `82.분배일`, `83.특허비용`, `84.중개수수료`, `86.발명자`, `88.산학협력단`, `88-1`, `88-2`, `88-3`
    """)
 
    col1, col2 = st.columns(2)
    with col1:
        master_file_tab2 = st.file_uploader(
            "📊 기술이전 총정리파일 (.xlsx)", type=['xlsx'], key="master2"
        )
    with col2:
        dist_files = st.file_uploader(
            "💰 기술료 분배 PDF (여러 건 가능)", type=['pdf'],
            accept_multiple_files=True, key="dist"
        )
 
    if st.button("🔄 분배 데이터 업데이트 시작", use_container_width=True, key="btn2"):
        if not master_file_tab2:
            st.error("⚠️ 기술이전 총정리파일을 업로드해 주세요!")
        elif not dist_files:
            st.error("⚠️ 분배 PDF를 최소 1개 이상 업로드해 주세요!")
        else:
            model_name = get_best_model()
 
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_master:
                tmp_master.write(master_file_tab2.read())
                master_path = tmp_master.name
 
            all_dist_data = []
            progress_bar = st.progress(0)
            status_text = st.empty()
 
            for i, d_file in enumerate(dist_files):
                status_text.info(f"⏳ [{i+1}/{len(dist_files)}] '{d_file.name}' 분석 중...")
 
                with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_d:
                    tmp_d.write(d_file.read())
                    d_path = tmp_d.name
 
                dist_data = extract_distribution_with_gemini(d_path, model_name)
                all_dist_data.append(dist_data)
 
                os.remove(d_path)
                progress_bar.progress((i + 1) / len(dist_files))
 
            status_text.info("📝 총정리파일 분배 열 업데이트 중...")
            updated_bytes, results = update_distribution_in_master(master_path, all_dist_data)
            os.remove(master_path)
 
            status_text.success("🎉 분배 데이터 업데이트 완료!")
 
            # 결과 표시
            st.markdown("### 📋 업데이트 결과")
            for r in results:
                if "✅" in r["상태"]:
                    st.success(f"연번 **{r['연번']}** — {r['상태']}")
                else:
                    st.error(f"연번 **{r['연번']}** — {r['상태']}")
 
            today = datetime.date.today().strftime("%Y%m%d")
            st.download_button(
                label="📥 분배 업데이트된 기술이전총정리파일 다운로드",
                data=updated_bytes,
                file_name=f"기술이전총정리_분배업데이트_{today}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
